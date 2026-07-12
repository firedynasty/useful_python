#!/usr/bin/env python3
"""Split an .xlsx file into one CSV per sheet.

Usage: xlsx_to_csvs.py <input.xlsx> [output_dir]

Creates a folder named <input>_csvs/ (or the specified output_dir)
containing one .csv file per sheet.
"""
import sys
import os
import csv
import openpyxl


def xlsx_to_csvs(xlsx_path, output_dir=None):
    if not os.path.isfile(xlsx_path):
        print(f"Error: file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    base = os.path.splitext(xlsx_path)[0]
    if output_dir is None:
        output_dir = base + "_csvs"

    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Sanitize sheet name for use as filename
        safe_name = sheet_name.replace("/", "_").replace("\\", "_").replace(":", "_")
        csv_path = os.path.join(output_dir, f"{safe_name}.csv")

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else cell for cell in row])

        print(f"  -> {csv_path}")

    wb.close()
    print(f"\nDone. {len(wb.sheetnames)} sheet(s) exported to {output_dir}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: xlsx_to_csvs.py <input.xlsx> [output_dir]", file=sys.stderr)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else None
    xlsx_to_csvs(xlsx_path, output_dir)
