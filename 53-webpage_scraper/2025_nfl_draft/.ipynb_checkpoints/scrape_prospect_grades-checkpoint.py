"""
Interactive NFL prospect grade scraper.

1. Shows you a player name to search
2. You paste it into Google in Chrome, navigate to their NFL.com prospect page
3. Press Enter and the script reads the grade from the open page
4. Repeat for each player

Requires Chrome running with:
  /Applications/Google\ Chrome.app/Contents/MacOS/Google\ Chrome \
    --remote-debugging-port=9222 --user-data-dir=/tmp/chrome_debug_profile

Usage:
  python scrape_prospect_grades.py
"""

import time
import re
import csv
import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup


def connect_to_chrome():
    options = Options()
    options.add_experimental_option("debuggerAddress", "127.0.0.1:9222")
    return webdriver.Chrome(options=options)


def convert_height(raw):
    if not raw:
        return ""
    val = str(int(float(raw))).zfill(4)
    feet, inches, eighths = int(val[0]), int(val[1:3]), int(val[3])
    if eighths == 0:
        return f"{feet}'{inches}\""
    return f"{feet}'{inches} {eighths}/8\""


def convert_measurement(raw):
    if not raw:
        return ""
    val = str(int(float(raw)))
    if len(val) < 3:
        return val
    inches, eighths = int(val[:-2]), int(val[-2])
    if eighths == 0:
        return f'{inches}"'
    return f'{inches} {eighths}/8"'


def fmt(val):
    if val is None or val == "":
        return ""
    return str(val)


def fmt_weight(val):
    if val is None or val == "":
        return ""
    return str(int(float(val)))


def norm_name(name):
    n = name.lower().strip()
    n = re.sub(r"\s+(jr\.?|sr\.?|ii|iii|iv|v)$", "", n)
    return re.sub(r"[.']", "", n)


def find_combine(player_name, combine):
    key = player_name.strip().lower()
    if key in combine:
        return combine[key]
    nk = norm_name(player_name)
    for k, v in combine.items():
        if norm_name(k) == nk:
            return v
    return None


def read_grade_from_current_page(driver):
    """Read prospect grade from whatever NFL.com prospect page is currently open."""
    html = driver.page_source
    soup = BeautifulSoup(html, "html.parser")
    lines = [l.strip() for l in soup.get_text(separator="\n").split("\n") if l.strip()]

    grade = None
    for i, line in enumerate(lines):
        if line == "Prospect Grade" and i + 1 < len(lines):
            grade = lines[i + 1]
            break
    return grade


def load_progress(progress_file):
    if not os.path.exists(progress_file):
        return {}
    grades = {}
    with open(progress_file, "r") as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) >= 2:
                grades[row[0]] = row[1]
    return grades


def save_progress(progress_file, grades):
    with open(progress_file, "w", newline="") as f:
        writer = csv.writer(f)
        for name, grade in grades.items():
            writer.writerow([name, grade])


def main():
    xlsx_path = os.path.expanduser(
        "~/Downloads/NFL Combine 2025 @JordanSportGuy Twitter.xlsx"
    )
    output_path = "2025_nfl_draft_with_combine.csv"
    progress_file = "grade_progress.csv"

    driver = connect_to_chrome()

    # 1. Scrape PFR draft data
    print("Scraping PFR draft data...")
    driver.get("https://www.pro-football-reference.com/years/2025/draft.htm")
    time.sleep(5)
    table_html = driver.execute_script("return document.getElementById('drafts').outerHTML;")
    soup = BeautifulSoup(table_html, "html.parser")
    draft_data = []
    for row in soup.find("tbody").find_all("tr"):
        if "thead" in row.get("class", []):
            continue
        rd = {
            cell.get("data-stat", ""): cell.get_text(strip=True)
            for cell in row.find_all(["th", "td"])
        }
        if rd.get("player"):
            draft_data.append(rd)

    # Filter to rounds 1-3
    draft_r1_3 = [d for d in draft_data if d.get("draft_round", "") in ("1", "2", "3")]
    print(f"Rounds 1-3: {len(draft_r1_3)} players\n")

    # 2. Load combine data
    combine = {}
    wb = openpyxl.load_workbook(xlsx_path)
    for sn in wb.sheetnames:
        ws = wb[sn]
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if not row[0]:
                continue
            combine[row[0].strip().lower()] = dict(zip(header, row))

    # 3. Load previous progress
    grades = load_progress(progress_file)
    already = sum(1 for d in draft_r1_3 if d["player"] in grades)
    if already:
        print(f"Resuming - {already} players already scraped\n")

    # 4. Interactive grade scraping
    print("=" * 60)
    print("INSTRUCTIONS:")
    print("  1. I'll show you a player name")
    print("  2. Search for them on Google in your Chrome window")
    print("  3. Click through to their NFL.com prospect page")
    print("  4. Come back here and press Enter")
    print("  Type 's' to skip, 'q' to quit and save")
    print("=" * 60)

    for i, d in enumerate(draft_r1_3):
        player = d["player"]

        if player in grades:
            print(f"[{i+1}/{len(draft_r1_3)}] {player} - already have: {grades[player]}")
            continue

        print(f"\n[{i+1}/{len(draft_r1_3)}] Search for: {player} nfl prospect")
        response = input("  Press Enter when on prospect page (s=skip, q=quit): ").strip().lower()

        if response == "q":
            print("Saving and quitting...")
            break
        elif response == "s":
            grades[player] = ""
            print("  Skipped")
            continue

        grade = read_grade_from_current_page(driver)
        grades[player] = grade if grade else ""
        print(f"  -> Grade: {grade if grade else 'not found'}")

        save_progress(progress_file, grades)

    # 5. Build merged output
    output_header = [
        "Round", "Pick", "Team", "Player", "Position", "Age", "College",
        "Prospect Grade",
        "Height", "Weight", "Hand Size", "Arm Length", "Wingspan",
        "40 Yard Dash", "10 Yard Split", "Vertical", "Broad",
        "3 Cone", "Shuttle", "Bench",
    ]

    output_rows = []
    for d in draft_r1_3:
        c = find_combine(d["player"], combine)
        row = [
            d.get("draft_round", ""), d.get("draft_pick", ""), d.get("team", ""),
            d.get("player", ""), d.get("pos", ""), d.get("age", ""),
            d.get("college_id", ""),
            grades.get(d["player"], ""),
            convert_height(fmt(c.get("HEIGHT"))) if c else "",
            fmt_weight(c.get("WEIGHT")) if c else "",
            convert_measurement(fmt(c.get("Hand Size "))) if c else "",
            convert_measurement(fmt(c.get("Arm Length"))) if c else "",
            convert_measurement(fmt(c.get("Wingspan"))) if c else "",
            fmt(c.get("40 Yard Dash")) if c else "",
            fmt(c.get("10 Yard Split")) if c else "",
            fmt(c.get("Vertical")) if c else "",
            fmt(c.get("Broad")) if c else "",
            fmt(c.get("3 Cone")) if c else "",
            fmt(c.get("Shuttle")) if c else "",
            fmt(c.get("Bench")) if c else "",
        ]
        output_rows.append(row)

    with open(output_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(output_header)
        writer.writerows(output_rows)

    graded = sum(1 for d in draft_r1_3 if grades.get(d["player"]))
    combined = sum(1 for d in draft_r1_3 if find_combine(d["player"], combine))
    print(f"\nSaved to {output_path}")
    print(f"  {len(draft_r1_3)} players, {graded} with grades, {combined} with combine data")

    driver.quit()


if __name__ == "__main__":
    main()
