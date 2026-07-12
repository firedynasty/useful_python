#!/usr/bin/env python3
"""
Turn a time-tracking CSV (date, time, activity) into an Excel workbook
with per-entry durations, a running work-hours tally, and a daily summary.

Usage:
    python time_tracker.py Tracking_work_time.csv
    python time_tracker.py Tracking_work_time.csv -o MyReport.xlsx
"""
import argparse
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATE_FORMATS = ('%m-%d-%y', '%m/%d/%Y', '%Y-%m-%d')


def parse_date(s):
    s = s.strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognized date format: {s!r}")


def load_csv(path):
    df = pd.read_csv(path)
    df.columns = ['date', 'time', 'activity']
    df['date'] = df['date'].apply(parse_date)
    df['time'] = pd.to_datetime(df['time'].str.strip(), format='%I:%M %p').dt.time
    df['activity'] = df['activity'].str.strip().str.lower()
    return df


def build_workbook(df, activities_to_sum=('work', 'break', 'lunch')):
    n = len(df)
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Log"

    header_fill = PatternFill('solid', start_color='305496')
    header_font = Font(bold=True, color='FFFFFF', name='Arial')
    font = Font(name='Arial')
    bold = Font(bold=True, name='Arial')
    border = Border(*(Side(style='thin', color='D9D9D9'),) * 4)

    headers = ['Date', 'Time', 'Activity', 'Start Datetime (helper)',
               'Duration to next entry (hrs)', 'Cumulative Work Hours (running, by date)']
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border

    for i, row in df.iterrows():
        r = i + 2
        ws.cell(row=r, column=1, value=row['date']).number_format = 'mm/dd/yyyy'
        ws.cell(row=r, column=2, value=row['time']).number_format = 'h:mm AM/PM'
        ws.cell(row=r, column=3, value=row['activity'])
        ws.cell(row=r, column=4, value=f'=A{r}+B{r}').number_format = 'mm/dd/yyyy h:mm AM/PM'
        if r < n + 1:
            ws.cell(row=r, column=5,
                     value=f'=IF(A{r+1}=A{r},(D{r+1}-D{r})*24,"open")')
        work_add = f'IF(AND(C{r}="work",ISNUMBER(E{r})),E{r},0)'
        cum = f'={work_add}' if r == 2 else f'=IF(A{r}=A{r-1},F{r-1},0)+{work_add}'
        ws.cell(row=r, column=6, value=cum).number_format = '0.00'
        for c in range(1, 7):
            ws.cell(row=r, column=c).font = font
            ws.cell(row=r, column=c).border = border

    ws.cell(row=n + 1, column=5, value='open (no next entry)').font = font
    for i, w in enumerate([12, 12, 16, 24, 26, 30], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'

    # Daily Summary
    ws2 = wb.create_sheet('Daily Summary')
    headers2 = ['Date'] + [f'Total {a.title()} Hours' for a in activities_to_sum]
    for c, h in enumerate(headers2, start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = border

    unique_dates = sorted(df['date'].unique())
    for i, d in enumerate(unique_dates):
        r = i + 2
        c1 = ws2.cell(row=r, column=1, value=d)
        c1.number_format, c1.font, c1.border = 'mm/dd/yyyy', font, border
        for col, activity in enumerate(activities_to_sum, start=2):
            f = f"=SUMIFS('Time Log'!E:E,'Time Log'!A:A,A{r},'Time Log'!C:C,\"{activity}\")"
            cell = ws2.cell(row=r, column=col, value=f)
            cell.number_format, cell.font, cell.border = '0.00', font, border

    total_r = len(unique_dates) + 2
    ws2.cell(row=total_r, column=1, value='Total').font = bold
    for col in range(2, 2 + len(activities_to_sum)):
        letter = get_column_letter(col)
        cell = ws2.cell(row=total_r, column=col, value=f'=SUM({letter}2:{letter}{total_r-1})')
        cell.number_format, cell.font = '0.00', bold

    for i, w in enumerate([14] + [18] * len(activities_to_sum), start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'

    return wb


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument('csv_path')
    ap.add_argument('-o', '--output', default=None)
    args = ap.parse_args()

    df = load_csv(args.csv_path)
    wb = build_workbook(df)
    out = args.output or args.csv_path.rsplit('.', 1)[0] + '.xlsx'
    wb.save(out)
    print(f"Saved {out}")


if __name__ == '__main__':
    main()
